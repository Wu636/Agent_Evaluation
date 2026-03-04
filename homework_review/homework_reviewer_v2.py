import asyncio
import json
import json.decoder
import os
import tempfile
import time
import uuid
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional

import requests
from dotenv import load_dotenv

# 本地解析模块（跳过云端 API）
try:
    from local_parser import parse_word_to_text_input
    LOCAL_PARSER_AVAILABLE = True
except ImportError:
    LOCAL_PARSER_AVAILABLE = False

# LLM 答案校验模块
try:
    from llm_answer_corrector import async_correct_answers_with_llm
    LLM_CORRECTOR_AVAILABLE = True
except ImportError:
    LLM_CORRECTOR_AVAILABLE = False


def load_env_config():
    """
    加载.env配置文件，优先加载当前目录下的.env文件
    如果当前目录没有，则加载上级目录的.env文件
    """
    current_dir = Path(__file__).parent

    # 优先尝试加载当前目录下的.env文件
    local_env = current_dir / '.env'
    if local_env.exists():
        load_dotenv(local_env)
        print(f"✅ 从本地目录加载.env配置: {local_env}")
        return local_env

    # 如果当前目录没有，尝试加载上级目录的.env文件
    parent_env = current_dir.parent / '.env'
    if parent_env.exists():
        load_dotenv(parent_env)
        print(f"✅ 从上级目录加载.env配置: {parent_env}")
        return parent_env

    # 如果都没有找到，尝试从当前工作目录加载
    cwd_env = Path.cwd() / '.env'
    if cwd_env.exists():
        load_dotenv(cwd_env)
        print(f"✅ 从工作目录加载.env配置: {cwd_env}")
        return cwd_env

    raise FileNotFoundError("未找到.env配置文件，请在当前目录或上级目录创建.env文件")


def safe_json_loads(value):
    if isinstance(value, (dict, list)):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except json.JSONDecodeError:
            return None
    return None


def extract_writing_requirement(detail: dict) -> str:
    business_config = safe_json_loads(detail.get("businessConfig"))
    writing_requirement = ""
    if isinstance(business_config, dict):
        composition = business_config.get("compositionRequirement") or {}
        writing_requirement = composition.get("writingRequirement") or ""
        if not writing_requirement:
            requirement_file = composition.get("requirementFile") or {}
            writing_requirement = requirement_file.get("content") or ""
    if not writing_requirement:
        writing_requirement = detail.get("desc") or ""
    return writing_requirement


def fetch_instance_details(instance_nid: str):
    """通过 agent/details 接口获取作业信息"""
    url = "https://cloudapi.polymas.com/agents/v1/agent/details"

    authorization = os.getenv('AUTHORIZATION')
    cookie = os.getenv('COOKIE')
    if not authorization:
        print("❌ 未找到AUTHORIZATION环境变量，请在.env文件中配置AUTHORIZATION")
        return None
    if not cookie:
        print("❌ 未找到COOKIE环境变量，请在.env文件中配置COOKIE")
        return None

    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Authorization": authorization,
        "Cookie": cookie,
    }

    payload = {
        "instanceIds": [instance_nid],
        "needToToolSchema": False
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            data=json.dumps(payload, ensure_ascii=False).encode('utf-8')
        )
        result = response.json()
    except json.decoder.JSONDecodeError:
        print(f"❌ 获取作业信息失败，状态码：{response.status_code}")
        print("响应内容（非JSON格式，可能为服务端错误页）：", response.text)
        return None
    except Exception as e:
        print(f"❌ 获取作业信息异常：{str(e)}")
        return None

    if not result.get('success'):
        print(f"❌ 获取作业信息失败：{result.get('msg')}")
        return None

    instance_details = result.get('data', {}).get('instanceDetails', [])
    if not instance_details:
        print("❌ 获取作业信息失败：instanceDetails 为空")
        return None

    detail = instance_details[0] or {}
    user_id = detail.get('userId')
    agent_id = detail.get('agentNid') or detail.get('agentId')
    if not user_id:
        print("❌ 获取作业信息失败：响应中未找到 userId")
        return None
    if not agent_id:
        print("❌ 获取作业信息失败：响应中未找到 agentNid")
        return None

    writing_requirement = extract_writing_requirement(detail)

    return {
        "user_id": user_id,
        "agent_id": agent_id,
        "instance_name": detail.get("instanceName", ""),
        "desc": detail.get("desc", ""),
        "writing_requirement": writing_requirement,
        "version": detail.get("version") or 2,
    }


def ensure_instance_context():
    """通过接口获取实例信息（仅当前进程使用，不写回.env）"""
    instance_nid = os.getenv('INSTANCE_NID', '').strip().strip('"').strip("'")
    if not instance_nid:
        print("❌ 未找到INSTANCE_NID环境变量，请在.env文件中配置INSTANCE_NID")
        return None

    details = fetch_instance_details(instance_nid)
    if not details:
        return None

    user_id = details.get("user_id") or os.getenv("USER_ID", "").strip().strip('"').strip("'")
    agent_id = details.get("agent_id") or os.getenv("AGENT_ID", "").strip().strip('"').strip("'")
    if not user_id:
        print("❌ 未获取到 userId，请检查 INSTANCE_NID 是否正确")
        return None
    if not agent_id:
        print("❌ 未获取到 agentId，请检查 INSTANCE_NID 是否正确")
        return None

    print(f"✅ 已获取USER_ID: {user_id}")
    print(f"✅ 已获取AGENT_ID: {agent_id}")

    return {
        "instance_nid": instance_nid,
        "user_id": user_id,
        "agent_id": agent_id,
        "writing_requirement": details.get("writing_requirement", ""),
        "version": details.get("version") or 2,
        "instance_name": details.get("instance_name", ""),
        "desc": details.get("desc", ""),
    }


def upload_file(file_path):
    """
    上传文件到服务器

    Args:
        file_path: 本地文件路径

    Returns:
        dict: 包含 fileName 和 fileUrl 的字典，如果上传失败返回 None
    """
    url = "https://cloudapi.polymas.com/basic-resource/file/upload"

    # 生成唯一标识码
    identify_code = str(uuid.uuid4())

    try:
        # 打开文件
        with open(file_path, 'rb') as f:
            # 获取文件名和大小
            file_name = os.path.basename(file_path)
            file_size = os.path.getsize(file_path)

            # 根据文件扩展名判断 MIME 类型
            file_ext = os.path.splitext(file_name)[1].lower()
            mime_types = {
                '.png': 'image/png',
                '.jpg': 'image/jpeg',
                '.jpeg': 'image/jpeg',
                '.gif': 'image/gif',
                '.pdf': 'application/pdf',
                '.doc': 'application/msword',
                '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            }
            mime_type = mime_types.get(file_ext, 'application/octet-stream')

            # 准备表单数据
            files = {
                'file': (file_name, f, mime_type)
            }

            data = {
                'identifyCode': identify_code,
                'name': file_name,
                'chunk': '0',
                'chunks': '1',
                'size': str(file_size)
            }

            # 从环境变量中读取配置
            authorization = os.getenv('AUTHORIZATION')
            cookie = os.getenv('COOKIE')

            if not authorization:
                raise ValueError("未找到AUTHORIZATION环境变量，请在.env文件中配置AUTHORIZATION")
            if not cookie:
                raise ValueError("未找到COOKIE环境变量，请在.env文件中配置COOKIE")

            headers = {
                'Authorization': authorization,
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/142.0.0.0 Safari/537.36',
                'Cookie': cookie
            }

            # 发送请求
            print(f"⏳ 正在上传文件: {file_name}")
            response = requests.post(url, headers=headers, data=data, files=files)
            result = response.json()

            if result.get('success'):
                data = result.get('data', {})
                file_url = data.get('ossUrl')
                print(f"✅ 文件上传成功: {file_name}")
                return {
                    'fileName': file_name,
                    'fileUrl': file_url
                }
            else:
                print(f"❌ 文件上传失败: {file_name}, 错误信息: {result.get('msg')}")
                return None

    except FileNotFoundError:
        print(f"❌ 文件不存在: {file_path}")
        return None
    except Exception as e:
        print(f"❌ 上传文件时发生错误: {file_path}, 错误: {str(e)}")
        return None


def is_success_response(result: dict) -> bool:
    if not isinstance(result, dict):
        return False
    if "success" in result:
        return bool(result.get("success"))
    return result.get("code") == 200


def fetch_task_result(task_id: str, context: dict):
    """轮询获取任务结果"""
    url = "https://cloudapi.polymas.com/agents/v1/get/task"

    authorization = os.getenv('AUTHORIZATION')
    cookie = os.getenv('COOKIE')
    if not authorization:
        return False, {"error": "未找到AUTHORIZATION环境变量，请在.env文件中配置AUTHORIZATION"}
    if not cookie:
        return False, {"error": "未找到COOKIE环境变量，请在.env文件中配置COOKIE"}

    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Authorization": authorization,
        "Cookie": cookie
    }

    payload = {
        "taskId": task_id,
        "metadata": {
            "instanceNid": context.get("instance_nid", "")
        }
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            data=json.dumps(payload, ensure_ascii=False).encode('utf-8')
        )
        try:
            result = response.json()
        except json.decoder.JSONDecodeError:
            return False, {
                "status_code": response.status_code,
                "text": response.text
            }

        return is_success_response(result), result

    except Exception as e:
        return False, {"error": str(e)}


def poll_task_until_complete(task_id: str, context: dict, interval_seconds: int = 2, timeout_seconds: int = 300):
    start_time = time.monotonic()
    last_result = None

    while True:
        success, result = fetch_task_result(task_id, context)
        last_result = result

        if success and isinstance(result, dict):
            data = result.get("data") or {}
            if isinstance(data, dict):
                if data.get("artifacts"):
                    return True, result
                status = data.get("status") or {}
                state = status.get("state")
                if state == "completed":
                    return True, result
                if state in {"failed", "error", "cancelled"}:
                    return False, result
        else:
            return False, result

        if time.monotonic() - start_time >= timeout_seconds:
            return False, {
                "error": "任务超时",
                "taskId": task_id,
                "last_response": last_result
            }

        time.sleep(interval_seconds)


def normalize_text_input(raw_data) -> Optional[str]:
    parsed = raw_data
    if isinstance(raw_data, str):
        try:
            parsed = json.loads(raw_data)
        except json.JSONDecodeError:
            return raw_data

    if isinstance(parsed, dict) and "content" in parsed:
        items = parsed.get("content") or []
        return json.dumps(_normalize_content_items(items), ensure_ascii=False)

    if isinstance(parsed, list):
        return json.dumps(_normalize_content_items(parsed), ensure_ascii=False)

    if isinstance(parsed, dict):
        return json.dumps(parsed, ensure_ascii=False)

    return str(parsed) if parsed is not None else None


def _normalize_content_items(items) -> list:
    normalized = []
    if not isinstance(items, list):
        return normalized

    for item in items:
        if isinstance(item, dict):
            normalized.append(
                {
                    "itemId": item.get("itemId") or item.get("item_id") or "",
                    "itemName": item.get("itemName") or item.get("item_name") or "",
                    "stuAnswerContent": item.get("stuAnswerContent")
                    or item.get("stu_answer_content")
                    or item.get("content")
                    or "",
                }
            )
        else:
            normalized.append(
                {
                    "itemId": "",
                    "itemName": "",
                    "stuAnswerContent": str(item),
                }
            )

    return normalized


def homework_file_analysis(file_info: dict, context: dict):
    """调用 homeworkFileAnalysis 接口解析作业文件"""
    url = "https://cloudapi.polymas.com/agents/v1/file/homeworkFileAnalysis"

    authorization = os.getenv('AUTHORIZATION')
    cookie = os.getenv('COOKIE')
    if not authorization:
        return False, {"error": "未找到AUTHORIZATION环境变量，请在.env文件中配置AUTHORIZATION"}, None
    if not cookie:
        return False, {"error": "未找到COOKIE环境变量，请在.env文件中配置COOKIE"}, None

    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Authorization": authorization,
        "Cookie": cookie
    }

    payload = {
        "agentId": context.get("agent_id", ""),
        "instanceNid": context.get("instance_nid", ""),
        "userNid": context.get("user_id", ""),
        "version": context.get("version") or 2,
        "writingRequirement": context.get("writing_requirement", ""),
        "activeMode": "upload",
        "editorContent": "",
        "fileList": [file_info],
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            data=json.dumps(payload, ensure_ascii=False).encode('utf-8')
        )

        try:
            result = response.json()
        except json.decoder.JSONDecodeError:
            return False, {
                "status_code": response.status_code,
                "text": response.text
            }, None

        if not is_success_response(result):
            return False, result, None

        text_input = normalize_text_input(result.get("data"))
        if not text_input:
            return False, {"error": "解析成功但未提取到可用的 textInput", "response": result}, None

        return True, result, text_input

    except Exception as e:
        return False, {"error": str(e)}, None


def execute_agent_text(text_input: str, context: dict):
    """调用 agent API 执行作业批改（TEXT_INPUT）"""
    url = "https://cloudapi.polymas.com/agents/v1/execute/agent"

    authorization = os.getenv('AUTHORIZATION')
    cookie = os.getenv('COOKIE')
    if not authorization:
        return False, {"error": "未找到AUTHORIZATION环境变量，请在.env文件中配置AUTHORIZATION"}
    if not cookie:
        return False, {"error": "未找到COOKIE环境变量，请在.env文件中配置COOKIE"}

    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Authorization": authorization,
        "Cookie": cookie
    }

    user_id = context.get("user_id") or os.getenv("USER_ID", "")
    instance_nid = context.get("instance_nid") or os.getenv("INSTANCE_NID", "")
    if not user_id:
        return False, {"error": "未获取到userId，请检查INSTANCE_NID"}
    if not instance_nid:
        return False, {"error": "未获取到instanceNid，请检查INSTANCE_NID"}

    if not isinstance(text_input, str):
        text_input = json.dumps(text_input, ensure_ascii=False)

    payload = {
        "metadata": {
            "dimension": "NONE",
            "instanceNid": instance_nid,
            "userIds": [user_id],
            "version": context.get("version") or 2,
            "async": True
        },
        "sendParams": {
            "message": {
                "kind": "message",
                "parts": [
                    {
                        "kind": "data",
                        "data": {
                            "writingRequirement": context.get("writing_requirement", ""),
                            "fileList": None,
                            "textInput": text_input,
                            "submitType": "TEXT_INPUT"
                        }
                    }
                ]
            }
        }
    }

    try:
        response = requests.post(
            url,
            headers=headers,
            data=json.dumps(payload, ensure_ascii=False).encode('utf-8')
        )

        try:
            result = response.json()
        except json.decoder.JSONDecodeError:
            return False, {
                "status_code": response.status_code,
                "text": response.text
            }

        return is_success_response(result), result

    except Exception as e:
        return False, {"error": str(e)}


def execute_agent_text_with_poll(text_input: str, context: dict, interval_seconds: int = 2, timeout_seconds: int = 300):
    success, result = execute_agent_text(text_input, context)
    if not success:
        return False, result

    data = result.get("data") if isinstance(result, dict) else None
    if isinstance(data, dict) and data.get("kind") == "task":
        task_id = data.get("id")
        if not task_id:
            return False, {"error": "未获取到taskId", "response": result}
        return poll_task_until_complete(task_id, context, interval_seconds, timeout_seconds)

    return success, result


def normalize_input_path(path_str: str) -> Path:
    """规范化用户输入路径"""
    path_str = path_str.strip().strip('"').strip("'")
    path = Path(path_str).expanduser()
    if not path.is_absolute():
        path = Path.cwd() / path
    return path


def collect_files_from_folder(folder_path: Path):
    """从文件夹中收集文件（忽略隐藏文件）"""
    if not folder_path.exists() or not folder_path.is_dir():
        return []
    return sorted([p for p in folder_path.iterdir() if p.is_file() and not p.name.startswith('.')])


def save_analysis_result(output_dir: Path, file_info: dict, result: dict, text_input: str):
    """保存 homeworkFileAnalysis 结果"""
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "analysis.json"
    payload = {
        "fileName": file_info.get("fileName"),
        "fileUrl": file_info.get("fileUrl"),
        "savedAt": datetime.now().isoformat(timespec="seconds"),
        "textInput": text_input,
        "response": result
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return output_path


def save_result(output_dir: Path, file_info: dict, attempt_index: int, attempt_total: int, success: bool, result: dict):
    """保存测评结果到文件"""
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"attempt_{attempt_index:02d}.json"
    payload = {
        "fileName": file_info.get("fileName"),
        "fileUrl": file_info.get("fileUrl"),
        "attempt": attempt_index,
        "attemptTotal": attempt_total,
        "success": success,
        "savedAt": datetime.now().isoformat(timespec="seconds"),
        "response": result
    }
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')
    return output_path


def load_pdf_generator():
    """动态加载 PDF 生成模块（避免作为包导入）"""
    primary = Path(__file__).parent / "generate_report.py"
    fallback = Path(__file__).parent / "test" / "generate_report.py"
    report_path = primary if primary.exists() else fallback
    if not report_path.exists():
        print(f"❌ 未找到PDF生成脚本: {primary} 或 {fallback}")
        return None

    import importlib.util

    spec = importlib.util.spec_from_file_location("generate_report", report_path)
    if not spec or not spec.loader:
        print("❌ 无法加载PDF生成脚本")
        return None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def generate_pdf_report(result: dict, output_path: Path):
    """根据接口返回结果生成PDF报告"""
    try:
        module = load_pdf_generator()
        if not module:
            return False
    except SystemExit as exc:
        print(f"❌ PDF依赖缺失：{exc}")
        return False
    except Exception as exc:
        print(f"❌ 加载PDF生成模块失败：{exc}")
        return False

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            suffix=".json",
            delete=False,
            encoding="utf-8",
            dir=output_path.parent
        ) as tmp:
            tmp.write(json.dumps(result, ensure_ascii=False, indent=2))
            tmp_path = Path(tmp.name)

        module.generate_pdf(str(tmp_path), str(output_path))
        tmp_path.unlink(missing_ok=True)
        return True
    except Exception as exc:
        print(f"❌ 生成PDF失败：{exc}")
        return False


def can_generate_pdf(result: dict) -> bool:
    if not isinstance(result, dict):
        return False
    data = result.get("data")
    return isinstance(data, dict) and "artifacts" in data


def save_output(output_dir: Path, file_info: dict, attempt_index: int, attempt_total: int, success: bool, result: dict, output_format: str):
    """根据输出格式保存结果"""
    if output_format == "pdf":
        if not success or not can_generate_pdf(result):
            reason = result.get("msg") or result.get("error")
            status_code = result.get("status_code")
            code = result.get("code")
            trace_id = result.get("traceId")
            extra = []
            if status_code:
                extra.append(f"status={status_code}")
            if code:
                extra.append(f"code={code}")
            if trace_id:
                extra.append(f"traceId={trace_id}")
            if reason:
                extra.append(f"reason={reason}")
            detail = f" ({', '.join(extra)})" if extra else ""
            print(f"⚠️ 结果失败或不支持PDF，跳过PDF生成: {file_info.get('fileName')}{detail}")
            return None
        output_path = output_dir / f"attempt_{attempt_index:02d}.pdf"
        ok = generate_pdf_report(result, output_path)
        return output_path if ok else None

    return save_result(output_dir, file_info, attempt_index, attempt_total, success, result)


def extract_category_from_name(name: str) -> str:
    """从题目名称提取类型，如 '单项选择题第1题' → '单项选择题'，'一、选择题第3题' → '选择题'"""
    import re
    if not name:
        return ""
    # 先去掉中文序号前缀：一、 二、 三、... 或 （一）（二）... 或 1. 2. ...
    cleaned = re.sub(r'^[\u4e00-\u9fa5\d]+[\u3001.\uff0e\s]\s*', '', name)
    cleaned = re.sub(r'^[\uff08(][\u4e00-\u9fa5\d]+[\uff09)]\s*', '', cleaned)
    # 匹配常见题型：单项选择题、判断题、简答题、论述题、案例分析题等
    match = re.match(r'^([\u4e00-\u9fa5]+题)', cleaned)
    if match:
        return match.group(1)
    # 处理特殊格式如 "案例分析题第1问"
    if "案例分析" in name:
        return "案例分析题"
    return ""


def calculate_category_scores(question_scores: list) -> dict:
    """按题目类型分组统计得分"""
    category_scores = {}  # {"单项选择题": {"score": 18, "total": 20}, ...}
    category_order = []  # 保持题型出现顺序
    
    for q in question_scores:
        if not isinstance(q, dict):
            continue
        name = q.get("name", "")
        category = extract_category_from_name(name)
        if not category:
            continue
        
        if category not in category_scores:
            category_scores[category] = {"score": 0, "total": 0}
            category_order.append(category)
        
        category_scores[category]["score"] += q.get("score", 0) or 0
        category_scores[category]["total"] += q.get("totalScore", 0) or 0
    
    return {"scores": category_scores, "order": category_order}


def extract_core_data(result: dict) -> Optional[dict]:
    """从接口响应中提取核心评分数据"""
    if not isinstance(result, dict):
        return None

    report_data = result.get("data", result)
    if isinstance(report_data, dict) and "artifacts" in report_data:
        artifacts = report_data.get("artifacts") or []
        if artifacts:
            parts = artifacts[0].get("parts") or []
            if parts:
                core_data = parts[0].get("data", {})
            else:
                core_data = {}
        else:
            core_data = {}
    else:
        core_data = report_data if isinstance(report_data, dict) else {}

    if not isinstance(core_data, dict):
        return None

    # 计算分类得分
    question_scores = core_data.get("questionScores", [])
    category_data = calculate_category_scores(question_scores)

    # 计算真实满分：优先用各题 totalScore 之和，否则用 API 返回的 fullMark
    api_full_mark = core_data.get("fullMark", 100)
    sum_of_totals = sum(
        (q.get("totalScore", 0) or 0) for q in question_scores if isinstance(q, dict)
    )
    # 如果能从题目中算出满分且大于0，使用计算值；否则回退到 API 值
    real_full_mark = sum_of_totals if sum_of_totals > 0 else api_full_mark

    return {
        "total_score": core_data.get("totalScore"),
        "full_mark": real_full_mark,
        "dimension_scores": core_data.get("dimensionScores", []),
        "category_scores": category_data.get("scores", {}),
        "category_order": category_data.get("order", []),
        "question_scores": question_scores,  # 原始逐题评分
    }


def resolve_summary_output_root(file_paths, output_root: Optional[Path]) -> Path:
    if output_root:
        return output_root
    parents = {path.parent.resolve() for path in file_paths}
    if len(parents) == 1:
        return next(iter(parents)) / "review_results"
    return Path.cwd() / "review_results"


def generate_excel_summary(summary_items, file_paths, attempts: int, output_root: Optional[Path]):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 未找到 openpyxl，无法生成Excel评分表。请先安装: pip install openpyxl")
        return None

    output_root = resolve_summary_output_root(file_paths, output_root)
    output_root.mkdir(parents=True, exist_ok=True)

    label_counts = defaultdict(int)
    label_by_path = {}
    for path in file_paths:
        key = str(path)
        if key not in label_by_path:
            base = path.stem
            label_counts[base] += 1
            label = base if label_counts[base] == 1 else f"{base}({label_counts[base]})"
            label_by_path[key] = label

    score_data = {}
    order_by_label = {}
    category_order_by_label = {}  # 分类题型顺序

    for item in summary_items:
        if not item or not item.get("success"):
            continue
        core_data = extract_core_data(item.get("result", {}))
        if not core_data:
            continue
        file_path = item.get("file_path", "")
        label = label_by_path.get(file_path, Path(file_path).stem if file_path else "未命名")
        entry = score_data.setdefault(
            label,
            {
                "full_mark": core_data.get("full_mark", 100),
                "total_scores": [None] * attempts,
                "dimensions": {},
                "categories": {},  # 新增：分类得分
            },
        )
        order = order_by_label.setdefault(label, [])
        category_order = category_order_by_label.setdefault(label, [])

        attempt_index = item.get("attempt_index", 0)
        if 1 <= attempt_index <= attempts:
            entry["total_scores"][attempt_index - 1] = core_data.get("total_score")

            # 处理分类得分
            category_scores = core_data.get("category_scores", {})
            for cat_name in core_data.get("category_order", []):
                if cat_name not in category_order:
                    category_order.append(cat_name)
                cat_data = category_scores.get(cat_name, {})
                cat_entry = entry["categories"].setdefault(
                    cat_name,
                    {"scores": [None] * attempts, "total": cat_data.get("total", 0)}
                )
                cat_entry["scores"][attempt_index - 1] = cat_data.get("score")
                # 更新满分值（取最大值）
                if cat_data.get("total", 0) > cat_entry["total"]:
                    cat_entry["total"] = cat_data.get("total", 0)

            for dim in core_data.get("dimension_scores", []):
                name = dim.get("evaluationDimension") or "未命名维度"
                if name not in order:
                    order.append(name)
                scores = entry["dimensions"].setdefault(name, [None] * attempts)
                scores[attempt_index - 1] = dim.get("dimensionScore")

    if not score_data:
        print("⚠️ 未获取到可用于生成评分表的结果")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "评分表"

    headers = ["档次/学生", "评价维度"] + [f"第{i}次" for i in range(1, attempts + 1)] + ["均值", "方差"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 2
    
    # 按等级顺序排序：优秀、良好、中等、合格、较差
    level_order = {"优秀": 1, "良好": 2, "中等": 3, "合格": 4, "较差": 5}
    
    def get_level_priority(path):
        label = label_by_path.get(str(path), "")
        for level, priority in level_order.items():
            if level in label:
                return priority
        return 999  # 未匹配的放最后
    
    sorted_paths = sorted(file_paths, key=get_level_priority)
    
    for path in sorted_paths:
        label = label_by_path.get(str(path))
        if label not in score_data:
            continue
        entry = score_data[label]
        order = order_by_label.get(label, [])
        category_order = category_order_by_label.get(label, [])

        full_mark = entry.get("full_mark", 100)
        try:
            full_mark_text = str(int(full_mark))
        except (TypeError, ValueError):
            full_mark_text = str(full_mark)
        total_label = f"总分（{full_mark_text}分）" if full_mark_text else "总分"

        ws.cell(row=row_idx, column=1).value = label
        ws.cell(row=row_idx, column=2).value = total_label
        total_scores = entry.get("total_scores", [])
        for idx, score in enumerate(total_scores, start=3):
            ws.cell(row=row_idx, column=idx).value = score
        # 添加均值和方差
        valid_scores = [s for s in total_scores if s is not None]
        if valid_scores:
            mean_val = sum(valid_scores) / len(valid_scores)
            variance_val = sum((x - mean_val) ** 2 for x in valid_scores) / len(valid_scores) if len(valid_scores) > 1 else 0
            ws.cell(row=row_idx, column=3 + attempts).value = round(mean_val, 2)
            ws.cell(row=row_idx, column=4 + attempts).value = round(variance_val, 2)
        row_idx += 1

        # 输出分类得分（在总分后）
        for cat_name in category_order:
            cat_entry = entry["categories"].get(cat_name, {})
            cat_total = cat_entry.get("total", 0)
            try:
                cat_total_text = str(int(cat_total))
            except (TypeError, ValueError):
                cat_total_text = str(cat_total)
            cat_label = f"{cat_name}（{cat_total_text}分）" if cat_total_text else cat_name
            
            ws.cell(row=row_idx, column=1).value = ""
            ws.cell(row=row_idx, column=2).value = cat_label
            cat_scores = cat_entry.get("scores", [None] * attempts)
            for idx, score in enumerate(cat_scores, start=3):
                ws.cell(row=row_idx, column=idx).value = score
            # 添加均值和方差
            valid_scores = [s for s in cat_scores if s is not None]
            if valid_scores:
                mean_val = sum(valid_scores) / len(valid_scores)
                variance_val = sum((x - mean_val) ** 2 for x in valid_scores) / len(valid_scores) if len(valid_scores) > 1 else 0
                ws.cell(row=row_idx, column=3 + attempts).value = round(mean_val, 2)
                ws.cell(row=row_idx, column=4 + attempts).value = round(variance_val, 2)
            row_idx += 1

        for dim_name in order:
            ws.cell(row=row_idx, column=1).value = ""
            ws.cell(row=row_idx, column=2).value = dim_name
            scores = entry["dimensions"].get(dim_name, [None] * attempts)
            for idx, score in enumerate(scores, start=3):
                ws.cell(row=row_idx, column=idx).value = score
            # 添加均值和方差
            valid_scores = [s for s in scores if s is not None]
            if valid_scores:
                mean_val = sum(valid_scores) / len(valid_scores)
                variance_val = sum((x - mean_val) ** 2 for x in valid_scores) / len(valid_scores) if len(valid_scores) > 1 else 0
                ws.cell(row=row_idx, column=3 + attempts).value = round(mean_val, 2)
                ws.cell(row=row_idx, column=4 + attempts).value = round(variance_val, 2)
            row_idx += 1

        row_idx += 1

    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 18 if col_idx <= 2 else 12

    output_path = output_root / "评分表.xlsx"
    wb.save(output_path)
    print(f"✅ 评分表已生成: {output_path}")
    return output_path


async def async_upload_file(file_path: str, semaphore: asyncio.Semaphore):
    async with semaphore:
        return await asyncio.to_thread(upload_file, file_path)


async def async_homework_analysis(file_info: dict, context: dict, semaphore: asyncio.Semaphore):
    async with semaphore:
        return await asyncio.to_thread(homework_file_analysis, file_info, context)


async def async_execute_agent_text(text_input: str, context: dict, semaphore: asyncio.Semaphore):
    async with semaphore:
        return await asyncio.to_thread(execute_agent_text_with_poll, text_input, context)


async def evaluate_and_save(file_path: Path, file_info: dict, text_input: str, context: dict, output_dir: Path, attempt_index: int, attempt_total: int, output_format: str, semaphore: asyncio.Semaphore):
    print(f"⏳ 批改中: {file_info['fileName']} ({attempt_index}/{attempt_total})")
    
    # 批改重试机制（增强版 - 5次重试 + 指数退避）
    max_retries = 5
    base_delay = 3  # 基础延迟秒数
    success = False
    result = None
    
    for retry in range(max_retries):
        success, result = await async_execute_agent_text(text_input, context, semaphore)
        
        if success:
            break
        
        # 检查是否为可重试的错误
        error_msg = str(result.get("error", "")) if isinstance(result, dict) else str(result)
        error_msg_lower = error_msg.lower()
        is_retryable = any(keyword in error_msg for keyword in [
            "SSLError", "ConnectionError", "TimeoutError", "Max retries exceeded",
            "EOF occurred", "Connection reset", "Connection refused",
            "BadStatusLine", "RemoteDisconnected", "BrokenPipeError",
            "ChunkedEncodingError", "IncompleteRead",
        ]) or any(keyword in error_msg_lower for keyword in [
            "rate limit", "too many requests", "429",
            "500", "502", "503", "504",
            "internal server error", "bad gateway", "service unavailable", "gateway timeout",
            "network", "timeout", "connection", "reset", "broken pipe",
        ])
        
        if is_retryable and retry < max_retries - 1:
            # 指数退避：3s, 6s, 12s, 24s
            delay = base_delay * (2 ** retry)
            print(f"🔄 批改重试 ({retry + 1}/{max_retries - 1}): {file_info['fileName']} ({attempt_index}/{attempt_total}) - 等待{delay}s")
            print(f"   原因: {error_msg[:200]}")
            await asyncio.sleep(delay)
        else:
            if retry > 0:
                print(f"❌ 重试{retry}次后仍失败: {file_info['fileName']} ({attempt_index}/{attempt_total})")
            break
    
    output_path = save_output(output_dir, file_info, attempt_index, attempt_total, success, result, output_format)
    if output_path:
        print(f"✅ 完成: {file_info['fileName']} ({attempt_index}/{attempt_total}) -> {output_path}")
    else:
        print(f"⚠️ 完成: {file_info['fileName']} ({attempt_index}/{attempt_total}) -> 未生成文件")
    return {
        "file_path": str(file_path),
        "attempt_index": attempt_index,
        "attempt_total": attempt_total,
        "success": success,
        "result": result,
    }


async def run_batch(file_paths, attempts: int, context: dict, output_root: Optional[Path], output_format: str, max_concurrency: int = 5, local_parse: bool = False, skip_llm_files: str = None, file_groups: str = None):
    semaphore = asyncio.Semaphore(max_concurrency)

    # 解析需要跳过 LLM 校验的文件名列表
    skip_llm_set: set = set()
    if skip_llm_files:
        try:
            parsed = json.loads(skip_llm_files)
            if isinstance(parsed, list):
                skip_llm_set = set(parsed)
                print(f"ℹ️ 以下文件将跳过 LLM 校验: {', '.join(skip_llm_set)}")
        except (json.JSONDecodeError, TypeError):
            pass

    # 解析文件分组信息
    groups_map: dict = {}  # group_name -> [filename, ...]
    if file_groups:
        try:
            groups_map = json.loads(file_groups)
            if not isinstance(groups_map, dict):
                groups_map = {}
            else:
                for gname, fnames in groups_map.items():
                    print(f"📁 文件分组「{gname}」: {', '.join(fnames)}")
        except (json.JSONDecodeError, TypeError):
            groups_map = {}

    upload_tasks = [async_upload_file(str(path), semaphore) for path in file_paths]
    upload_results = await asyncio.gather(*upload_tasks)

    file_infos = []
    for path, result in zip(file_paths, upload_results):
        if result:
            file_infos.append((path, result))

    if not file_infos:
        print("\n❌ 没有成功上传的文件，无法执行批改")
        return

    print(f"\n✅ 成功上传 {len(file_infos)} 个文件，共 {len(file_paths)} 个")

    # 解析文件
    prepared_files = []
    
    if local_parse and LOCAL_PARSER_AVAILABLE:
        # 本地解析模式（跳过云端 API）
        print("\n📝 使用本地解析模式...")
        for path, file_info in file_infos:
            try:
                text_input = parse_word_to_text_input(path)
                file_root = output_root if output_root else (path.parent / "review_results")
                file_output_dir = file_root / path.stem
                file_output_dir.mkdir(parents=True, exist_ok=True)
                
                # 保存本地解析结果
                analysis_data = {
                    "fileName": file_info.get("fileName"),
                    "fileUrl": file_info.get("fileUrl"),
                    "savedAt": datetime.now().isoformat(),
                    "parseMode": "local",
                    "textInput": text_input
                }
                analysis_path = file_output_dir / "analysis.json"
                analysis_path.write_text(json.dumps(analysis_data, ensure_ascii=False, indent=2), encoding="utf-8")
                print(f"✅ 本地解析完成: {file_info.get('fileName')}")
                prepared_files.append((path, file_info, text_input, file_output_dir))
            except Exception as e:
                print(f"❌ 本地解析失败: {file_info.get('fileName')} ({e})")
    else:
        # 云端解析模式（带重试机制）
        max_parse_retries = 3
        parse_retry_delay = 2  # 重试间隔秒数
        
        for path, file_info in file_infos:
            success = False
            text_input = None
            analysis_result = None
            
            for retry in range(max_parse_retries):
                if retry > 0:
                    print(f"🔄 重试解析 ({retry}/{max_parse_retries-1}): {file_info.get('fileName')}")
                    await asyncio.sleep(parse_retry_delay)
                
                success, analysis_result, text_input = await async_homework_analysis(file_info, context, semaphore)
                
                if success and text_input:
                    break
            
            if not success or not text_input:
                reason = "解析失败"
                if isinstance(analysis_result, dict):
                    reason = analysis_result.get("msg") or analysis_result.get("error") or reason
                print(f"❌ 解析失败 (已重试{max_parse_retries-1}次): {file_info.get('fileName')} ({reason})")
                continue

            file_root = output_root if output_root else (path.parent / "review_results")
            file_output_dir = file_root / path.stem
            file_output_dir.mkdir(parents=True, exist_ok=True)
            analysis_path = save_analysis_result(file_output_dir, file_info, analysis_result, text_input)
            print(f"✅ 解析完成: {file_info.get('fileName')} -> {analysis_path}")

            # LLM 校验补充空白答案（用户标记跳过的文件不走 LLM 校验）
            file_name = file_info.get("fileName", "")
            should_skip_llm = file_name in skip_llm_set
            if should_skip_llm:
                print(f"ℹ️ 用户已标记跳过 LLM 校验: {file_name}")
            elif LLM_CORRECTOR_AVAILABLE:
                try:
                    text_input = await async_correct_answers_with_llm(path, text_input)
                except Exception as e:
                    print(f"⚠️ LLM 校验失败: {e}，继续使用原始解析结果")

            prepared_files.append((path, file_info, text_input, file_output_dir))

    if not prepared_files:
        print("\n❌ 没有成功解析的文件，无法执行批改")
        return

    # ── 分组合并逻辑 ──
    # 如果有分组信息，将同组文件的 text_input 合并为一份
    eval_items = []  # (label_path, file_info, text_input, output_dir) — 用于评测

    if groups_map:
        # 建立 filename -> prepared_file 的映射
        name_to_prepared = {}
        for item in prepared_files:
            path, file_info, text_input, file_output_dir = item
            fname = file_info.get("fileName", path.name)
            name_to_prepared[fname] = item

        grouped_names = set()
        for group_name, file_names in groups_map.items():
            members = [name_to_prepared[fn] for fn in file_names if fn in name_to_prepared]
            if len(members) < 2:
                # 不足2个文件的组不需要合并，按独立文件处理
                continue
            grouped_names.update(fn for fn in file_names if fn in name_to_prepared)

            # 合并 text_input
            combined_parts = []
            for path, file_info, text_input, _ in members:
                fname = file_info.get("fileName", path.name)
                combined_parts.append(f"--- 文件: {fname} ---\n{text_input}")
            merged_text = "\n\n".join(combined_parts)

            # 使用组名作为输出目录
            group_output_dir = (output_root if output_root else (members[0][0].parent / "review_results")) / group_name
            group_output_dir.mkdir(parents=True, exist_ok=True)

            # 使用第一个文件的 file_info，但修改 fileName 为组名
            group_file_info = dict(members[0][1])
            group_file_info["fileName"] = group_name
            group_file_info["_merged_files"] = [m[1].get("fileName", "") for m in members]

            print(f"📎 已合并分组「{group_name}」: {', '.join(group_file_info['_merged_files'])}")
            eval_items.append((members[0][0], group_file_info, merged_text, group_output_dir))

        # 未分组的文件独立评测
        for item in prepared_files:
            path, file_info, text_input, file_output_dir = item
            fname = file_info.get("fileName", path.name)
            if fname not in grouped_names:
                eval_items.append(item)
    else:
        eval_items = prepared_files

    tasks = []
    for path, file_info, text_input, file_output_dir in eval_items:
        for attempt_index in range(1, attempts + 1):
            tasks.append(
                evaluate_and_save(
                    path,
                    file_info,
                    text_input,
                    context,
                    file_output_dir,
                    attempt_index,
                    attempts,
                    output_format,
                    semaphore
                )
            )

    results = await asyncio.gather(*tasks)
    success_count = sum(1 for item in results if item and item.get("success"))
    print(f"\n✅ 已完成 {len(results)} 次测评（成功 {success_count}）")
    generate_excel_summary(results, [item[0] for item in eval_items], attempts, output_root)

    return {
        "results": results,
        "prepared_files": [str(item[0]) for item in eval_items],
        "output_root": str(output_root) if output_root else None,
        "attempts": attempts,
        "output_format": output_format,
        "success_count": success_count,
    }


def main():
    """主函数：处理用户交互和文件上传"""
    print("=" * 60)
    print("作业批改系统 - v2 (上传 -> 解析 -> 批改)")
    print("=" * 60)

    # 加载环境配置
    try:
        load_env_config()
    except FileNotFoundError as e:
        print(f"\n❌ {e}")
        return

    # 自动获取实例信息（不写回.env）
    context = ensure_instance_context()
    if not context:
        return

    instance_name = (context.get("instance_name") or "").strip()
    desc = (context.get("desc") or "").strip()
    if instance_name or desc:
        print("\n📌 作业信息：")
        if instance_name:
            print(f"名称: {instance_name}")
        if desc:
            print(f"描述: {desc}")

    # 选择上传方式
    print("\n请选择上传方式：")
    print("1) 单个/多个文件")
    print("2) 文件夹")
    upload_choice = input("请输入选项 (1/2，默认2): ").strip()

    file_paths = []
    output_root = None

    if upload_choice == "" or upload_choice == "2":
        folder_input = input("请输入文件夹路径: ").strip()
        if not folder_input:
            print("❌ 未输入文件夹路径")
            return
        folder_path = normalize_input_path(folder_input)
        file_paths = collect_files_from_folder(folder_path)
        if not file_paths:
            print("❌ 文件夹中未找到可上传的文件")
            return
        output_root = folder_path / "review_results"
    else:
        print("\n请输入要上传的文件路径（多个文件用逗号分隔）：")
        print("示例: /path/to/file1.png,/path/to/file2.jpeg")
        file_paths_input = input("文件路径: ").strip()
        if not file_paths_input:
            print("❌ 未输入文件路径")
            return
        file_paths = [normalize_input_path(path.strip()) for path in file_paths_input.split(',') if path.strip()]
        if not file_paths:
            print("❌ 未输入有效文件路径")
            return
        output_root = None

    # 询问测评次数
    attempts_input = input("每个文档需要测评几次？(默认5): ").strip()
    attempts = 5
    if attempts_input:
        try:
            attempts = int(attempts_input)
        except ValueError:
            print("❌ 测评次数必须为整数")
            return
        if attempts <= 0:
            print("❌ 测评次数必须大于0")
            return

    print("\n请选择报告格式：")
    print("1) JSON 报告（默认,生成速度快）")
    print("2) PDF 报告（需要完整评分结果）")
    report_choice = input("请输入选项 (1/2): ").strip().lower()
    output_format = "pdf" if report_choice in {"2", "pdf"} else "json"

    # 询问解析模式
    local_parse = False
    if LOCAL_PARSER_AVAILABLE:
        print("\n请选择解析模式：")
        print("1) 云端解析（默认，使用智慧树 API）")
        print("2) 本地解析（跳过云端 API，解析更稳定）")
        parse_choice = input("请输入选项 (1/2): ").strip()
        local_parse = parse_choice == "2"

    print(f"\n📂 共需要上传 {len(file_paths)} 个文件，每个文件测评 {attempts} 次")
    print(f"   输出格式: {output_format}，解析模式: {'本地' if local_parse else '云端'}\n")

    asyncio.run(run_batch(file_paths, attempts, context, output_root, output_format, local_parse=local_parse))


if __name__ == "__main__":
    main()
